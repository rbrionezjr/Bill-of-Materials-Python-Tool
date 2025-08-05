from arcgis.gis import GIS
import arcgis
from arcgis.geometry import filters
import arcpy
import json
from collections import defaultdict
import os
import sys
import openpyxl
from pathlib import Path


# Author - Ruben Brionez Jr
# Credits - Tiffany Rufo and the Omni Fiber GIS Department

# TODO: Add fail safe for inaccessible layers in portal
# TODO: Incorporate special crossings.
# TODO: Create a function to get a assumption on number of anchors based on strand features * 4
# TODO: Exclude features within an MDU Boundary
# TODO: Add a .lower() to anything that may be missed due to capitalization
# TODO: Add exclusion for 'Existing' features in Status
# TODO: Add a clipping function for strand and conduit for better accuracy OR defer to vendors for drafting
# TODO: Re-work the HHP calculations, revisit MDU, FDH HHPs and DNB HHPs - Use all address points? except in MDU?
# TODO: Messages should report back linear footages so people know things.

# Change Log 06-17-2024
# Version 1.4
""" - Added functionality to correctly select FDH layer name, either FDH_Boundary or FDH Boundary will work.
    - Changed AXL to XXL to match the materials list"""

# Change Log 04-22-2025
# Version 1.3
""" - Added AXL vault to structures function and returned it"""


# Change Log 04-22-2025
# Version 1.3
""" - Upgraded to version 1.3"""

# Change Log 04-15-2025
# Version 1.2
""" - Corrected the BOM Excel Template to correctly reference Thayer, LeeComm and Utilus and pull rates from the 
      RateCard tab.  
                    """

# Change Log 03-31-2025
# Version 1.2
""" - Upgraded from version 1.1 to 1.2
"""


# Change Log 03-28-2025
# Version 1.1
""" - Modified the cabinet size warning to clarify what may be causing the error.
    - Modified the warning about dividing by zero to clarify what calculation may be causing the issue. 
    - Corrected a bug where the Guys query was returning the correct count but message was incorrectly 
      referencing structures.
    - Added a lashing wire calculation.
    - Added special crossing footage as 4" conduit footage + 50.
"""


# Change Log 03-27-2025
# Version 1.1
""" - Modified the variable 'loop_type' to try and avoid the Nonetype error 
      if None gets passed to the .strip() function. This was near lines 947 and 948 of the script.   
    - Added clarification to the 4" Conduit Warning Message.
    - Added clarification to the negative PFA-2 Warnings.
    
      """

# Change Log 03-26-2025
# Version 1.1
""" - Added a calculation for special crossings by filtering the conduit size to 4 inches.
    - New Deployment method used. Send atbx, script, and template to a zip folder and upload to portal.
    - Added calculations for 1.25, 2, and 4 inch conduits.
    - Filling the special crossings cell with the sum of UG1 per 4 inch conduit.
    """

# Change Log 03-25-2025
# Version 1.1
""" - Added an escape from the export of the construction vendor and engineering vendor are left blank
    - Added a check for a negative value in PFA-2, this is usually the result of strand 
      from HL Draft left in the FDH Boundary.
    - Added the ability for users to select a different file directory for saving the excel export."""

# Change Log 03-24-2025
# Version 1.1
""" - Deployed as a Geoprocessing Tool instead of a web tool
    - Still downloadable from Portal
    - Updated the Splice Closure function to return the proper amount of variables (11)
    - Updated the AE percent to throw an error if dividing by zero
    - Updated script name to v1.1"""


# Connect using ArcGIS Pro’s active session (no credentials needed)
gis = GIS("pro")

# Add script directory to Python path
script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(script_dir)


arcpy.AddMessage("**** BOM Processing v1.4 - June 2025 ****\n"
                 "\n")


def get_one_drive_documents():
    user_profile = Path(os.environ["USERPROFILE"])
    # Looks for folders like "OneDrive" or "OneDrive - Omni Fiber LLC"
    for folder in user_profile.glob("OneDrive*/Documents"):
        return str(folder)
    # Fallback to classic Documents if OneDrive not found
    return str(user_profile / "Documents")


def export_to_excel(template_path, output_path, values_dict, construction_vendor_rate, design_vendor_rate):
    """Exports calculated values and fiber slack sums to specific cells in an existing Excel template."""

    try:
        # Load the existing Excel workbook
        wb = openpyxl.load_workbook(template_path)

        # Ensure the "RateCard" sheet exists before trying to write to it
        if "RateCard" in wb.sheetnames:
            rate_card_sheet = wb["RateCard"]
            rate_card_sheet["E2"] = construction_vendor_rate  # Write selected rate to cell E2
            # arcpy.AddMessage(f"► Construction Vendor '{construction_vendor_rate}' written to 'RateCard' sheet (E2).")
        else:
            arcpy.AddError("'RateCard' sheet not found in the Excel template.")

        # **Ensure the "RateCard_E" sheet exists and update design vendor rate**
        if "RateCard_E" in wb.sheetnames:
            rate_card_e_sheet = wb["RateCard_E"]
            rate_card_e_sheet["E2"] = design_vendor_rate  # ✅ Write design vendor to 'RateCard_E'!E2
            # arcpy.AddMessage(f"► Design Vendor '{design_vendor_rate}' written to 'RateCard_E'!E2.")
        else:
            arcpy.AddError("❌ 'RateCard_E' sheet not found in the Excel template.")

        # Select the main sheet
        if "Summary" in wb.sheetnames:
            summary_sheet = wb["Summary"]  # Explicitly selecting the main sheet
            summary_sheet["F5"] = design_vendor_rate  # ✅ Write engineering vendor to F5
            summary_sheet["F6"] = construction_vendor_rate  # ✅ Also write selection to F6
            summary_sheet["F7"] = datetime.now().strftime("%m-%d-%Y %H:%M:%S")  # MM-DD-YYYY format for timestamp
        else:
            arcpy.AddError("'Summary' sheet not found in the Excel template.")
            return  # Exit function if the sheet is missing

        # ✅ **New Mapping for Engineering Sheet**
        engineering_mapping = {
            "total_fiber_footage_ug_linear": ["F9"],
            "total_fiber_footage_ae_linear": ["F13"],
            "total_linear_footage": ["F7", "F15"],
            "e_epmrt_1": ["F19"]  # If it needs to be written to multiple places
        }

        # **Ensure the "Engineering" sheet exists and update values**
        if "Engineering" in wb.sheetnames:
            engineering_sheet = wb["Engineering"]

            for key, cell_list in engineering_mapping.items():
                if key in values_dict:  # ✅ Ensure key exists in values_dict
                    value = values_dict[key]  # ✅ Get the computed value

                    for cell in cell_list:  # ✅ Write to each mapped cell
                        engineering_sheet[cell] = value
                        # arcpy.AddMessage(f"📊 {key} written to 'Engineering' {cell}: {value}")
                else:
                    arcpy.AddWarning(f"⚠ {key} not found in values_dict. Skipping.")
        else:
            arcpy.AddError("❌ 'Engineering' sheet not found in the Excel template.")

        # **Base mapping of fixed values to specific Excel cells**
        cell_mapping = {
            "total_ug1ft": ['D66', 'D67', 'D131'],
            "total_ug2ft": ["D68"],
            "total_1in_conduit": ["D126"],
            "total_2in_conduit": ["D125"],
            "total_4in_conduit": ["D127"],
            "total_sp1": ["D93"],
            "total_sp2": ["D94"],
            "total_sp3_excluding_f1": ["D95"],
            "fiber_12": ["D108"],
            "fiber_24": ["D107"],
            "fiber_48": ["D106"],
            "fiber_96": ["D105"],
            "fiber_144": ["D104"],
            "fiber_288": ["D103"],
            "fiber_432": ["D110"],
            "total_heatshrink": ["D132"],
            "fp_count": ["D79", "D123"],
            "sv_count": ["D74", "D118"],
            "mv_count": ["D75", "D119"],
            "lv_count": ["D76", "D120"],
            "xl_count": ["D77", "D121"],
            "xsv_count": ["D73", "D117"],
            "nid_count": ["D142"],
            "axl_count": ["D122"],
            "coyote_count": ["D133"],
            "x17_count": ["D134"],
            "x22_count": ["D135"],
            "x28_count": ["D137"],
            "x19_count": ["D136"],
            "runt_count": ["D138"],
            "total_closure_count": ["D89"],
            "hanger_bracket": ["D140"],
            "offset_bracket": ["D141"],
            "lash_closure_count": ["D91"],
            "drop_count": ["C25"],
            "total_hhp_mdu": ["C24"],
            "total_strand_ftg": ['D49', "D50", "D113", "D114"],
            "est_total_miles": ['C28'],
            "ae_bom_miles": ['C29'],
            "ug_bom_miles": ['C30'],
            "percent_ae": ['C31'],
            "percent_ug": ['C32'],
            "total_hhp": ['C26'],
            "total_f1_miles": ["C35"],
            "total_f2_miles": ["C38"],
            "total_f2_ug": ["C40"],
            "total_f2_ae": ["C39"],
            "total_ae_ftg": ["C33"],
            "total_ug_ftg": ["C34"],
            "total_f1_ae": ["C36"],
            "total_f1_ug": ["C37"],
            "pfd_1": ['D71'],
            "passive_144": ["D149"],
            "passive_288": ["D150"],
            "passive_432": ["D151"],
            "passive_576": ["D152"],
            "ug_closure_count": ["D147", "D148"],
            "snowshoes": ["D92", "D113"],
            "conduit_couplers_1in": ["D128"],
            "conduit_couplers_2in": ["D129"],
            "conduit_couplers_4in": ["D130"],
            "pfa_2": ["D51"],
            "total_risers": ["D58", "D112"],
            "cab_id": ["F2"],
            "serv_area": ["F3"],
            "city_code": ["F4"],
            "total_strand_ftg_reareasment_y": ["D52"],
            "total_cabinets": ["C42"],
            "count_over_600ft": ["C41"],
            "average_calcfootage": ["C44"],
            "total_pole_count": ["C27"],
            "mr_filtered_pole_count": ["D59"],
            "active_cabinet_count": ["C43", "D86", "D87", "D78", "D122"],
            "grounded_poles": ["D57"],
            "tree_trimming": ["D53"],
            "total_ug1ft_reareasment_Y": ["D72"],
            "down_count": ["D54"],
            "dirt_count": ["D55"],
            "rock_count": ["D56"],
            "total_anchors": ["D144"],
            "uguard_adapter": ["D111"],
            "lashing_wire": ["D116"],
            "special_crossing": ["D70"]
        }

        # Write values to all specified cells
        for key, cell_list in cell_mapping.items():
            if key in values_dict:
                for cell in cell_list:  # Loop through multiple cell destinations
                    summary_sheet[cell] = values_dict[key]  # Write value to each mapped cell
                    # arcpy.AddMessage(f"✅ Writing {key} = {values_dict[key]} to {cell}")
            else:
                arcpy.AddMessage(f"⚠ {key} not found in values_dict. Skipping.")

        if "RateCard" in wb.sheetnames:
            wb["RateCard"].sheet_state = "hidden"  # ✅ Hide the sheet
            # arcpy.AddMessage("👀 'RateCard' sheet hidden.")

        if "RateCard_E" in wb.sheetnames:
            wb["RateCard_E"].sheet_state = "hidden"  # ✅ Hide the sheet
            # arcpy.AddMessage("👀 'RateCard_E' sheet hidden.")

        # Save the updated workbook
        arcpy.AddMessage("📁 Saving Excel file...")
        wb.save(output_path)
        arcpy.AddMessage(f"✅ Excel file successfully saved: {output_path}")

    except Exception as e:
        arcpy.AddError(f"❌ Error exporting to Excel: {e}")


def count_addresses(fdh_geometry):
    try:
        # Query Address Points within FDH Boundary
        address_layer_item = gis.content.get(address_master_id)
        if not address_layer_item:
            arcpy.AddError("❌ Address Master layer not found in Portal.")
            return 0, 0, 0, 0, 0, 0

        address_layer = address_layer_item.layers[0]  # Assuming first layer is correct
        address_query = address_layer.query(
            geometry_filter=filters.contains(fdh_geometry),
            out_fields="*",
            return_geometry=True
        )

        total_addresses = len(address_query.features)

        # Query MDU Polygon within FDH Boundary
        mdu_layer_item = gis.content.get(mdu_boundary_id)
        if not mdu_layer_item:
            arcpy.AddError("❌ MDU Boundary layer not found in Portal.")
            return total_addresses, 0, 0, 0, 0, 0, 0

        mdu_layer = mdu_layer_item.layers[0]
        mdu_query = mdu_layer.query(
            geometry_filter=filters.contains(fdh_geometry),
            out_fields="*",
            return_geometry=True
        )

        mdu_boundary_count = len(mdu_query.features)

        # Sum hhp_count values for MDU polygons within FDH boundary
        total_hhp_mdu = 0
        for mdu_feature in mdu_query.features:
            hhp_value_raw = mdu_feature.attributes.get('hhp_count', '0')
            try:
                hhp_value = int(hhp_value_raw)
            except (ValueError, TypeError):
                hhp_value = 0  # fallback if value is not convertible
            total_hhp_mdu += hhp_value

        # Query Do Not Build Polygon within FDH Boundary
        dnb_layer_item = gis.content.get(do_not_build_id)
        if not dnb_layer_item:
            arcpy.AddError("❌ Do Not Build Boundary layer not found in Portal.")
            return total_addresses, total_hhp_mdu, 0

        dnb_layer = dnb_layer_item.layers[0]
        dnb_query = dnb_layer.query(
            geometry_filter=filters.contains(fdh_geometry),
            out_fields="*",
            return_geometry=True
        )

        dnb_boundary_count = len(dnb_query.features)

        # Count address points in each Do Not Build polygon
        total_dnb_addresses = 0
        for dnb_feature in dnb_query.features:
            dnb_geometry = dnb_feature.geometry
            address_query_dnb = address_layer.query(
                geometry_filter=filters.contains(dnb_geometry),
                out_fields="*",
                return_geometry=True
            )
            total_dnb_addresses += len(address_query_dnb.features)

        # arcpy.AddMessage(f"🚫 Total Do Not Build Polygons: {dnb_boundary_count}")
        # arcpy.AddMessage(f"🚫 Total Addresses in Do Not Build Polygons: {total_dnb_addresses}")

        return total_addresses, total_hhp_mdu, total_dnb_addresses, mdu_boundary_count, dnb_boundary_count

    except Exception as e:
        arcpy.AddError(f"❌ Error processing address counts: {e}")
        return 0, 0, 0, 0, 0, 0  # Ensure function always returns three values


def fdh_boundary_selection_multiple(fdh_boundary_id):
    try:
        aprx = arcpy.mp.ArcGISProject("CURRENT")
        active_map = aprx.activeMap

        # Find the FDH_Boundary layer
        fdh_layer = next((layer for layer in active_map.listLayers()
                         if (layer.name == "FDH_Boundary" or layer.name == "FDH Boundary") and layer.isFeatureLayer), None)

        if not fdh_layer:
            arcpy.AddError("❌ FDH_Boundary layer not found in the active map.")
            return []

        # Get a count of the number of selected FDH Boundaries
        selected_count = int(arcpy.GetCount_management(fdh_layer)[0])
        if selected_count == 0:
            arcpy.AddError("❌ No FDH_Boundary features selected.")
            return []

        arcpy.AddMessage(f"🔍 Found {selected_count} selected FDH_Boundary features.")

        # Retrieve full layer from Portal
        fdh_item = gis.content.get(fdh_boundary_id)
        if not fdh_item:
            arcpy.AddError("❌ FDH_Boundary layer not found in Portal.")
            return []

        portal_layer = fdh_item.layers[0]

        # Get list of selected cab_ids from local selection
        cab_ids = []
        with arcpy.da.SearchCursor(fdh_layer, ["cab_id"]) as cursor:
            for row in cursor:
                cab_ids.append(row[0])

        arcpy.AddMessage(f"📋 Selected cab_ids: {cab_ids}")

        # Query portal layer for all selected cab_ids
        cab_ids_sql = f"cab_id IN ({','.join(repr(cid) for cid in cab_ids)})"
        query_result = portal_layer.query(
            where=cab_ids_sql,
            out_fields="*",
            return_geometry=True
        )

        if not query_result.features:
            arcpy.AddError("❌ No matching features found in portal layer.")
            return []

        selected_data = []
        for feature in query_result.features:
            selected_data.append({
                "object_id": feature.attributes.get("OBJECTID"),
                "geometry": feature.geometry,
                "cab_id": feature.attributes.get("cab_id"),
                "serv_area": feature.attributes.get("Serv_Area"),
                "city_code": feature.attributes.get("City_Code"),
                "hhp_count": feature.attributes.get("hhp_count"),
                "db_status": feature.attributes.get("DB_Status")
            })

        return selected_data

    except Exception as e:
        arcpy.AddError(f"❌ Error retrieving FDH boundaries: {e}")
        return []


def fdh_boundary_selection(fdh_boundary_id):
    try:
        cab_id = arcpy.GetParameterAsText(0).upper()

        if not cab_id:
            aprx = arcpy.mp.ArcGISProject("CURRENT")
            active_map = aprx.activeMap  # Get the active map

            # Find the FDH_Boundary layer in the active map
            fdh_layer = None
            for layer in active_map.listLayers():
                if (layer.name == "FDH_Boundary" or layer.name == "FDH Boundary") and layer.isFeatureLayer:
                    fdh_layer = layer
                    break

            if not fdh_layer:
                arcpy.AddError("❌ FDH_Boundary layer not found in the active map.")
                return None, None, None, None, None, None

            # Check if a feature is selected
            selected_count = int(arcpy.GetCount_management(fdh_layer)[0])
            if selected_count == 0:
                arcpy.AddError(
                    "❌ No FDH_Boundary feature selected. Please select one in the map or enter a cab_id.")
                return None, None, None, None, None, None

            # Retrieve cab_id from the selected feature
            with arcpy.da.SearchCursor(fdh_layer, ["cab_id"]) as cursor:
                for row in cursor:
                    cab_id = row[0]
                    break  # Take the first selected feature

            if not cab_id:
                arcpy.AddError("❌ The selected FDH_Boundary feature has no cab_id.")
                return None, None, None, None, None, None

            arcpy.AddMessage(f"► Selected FDH from map: {cab_id}")
            arcpy.SetParameter(0, cab_id)  # Sets the first parameter as the cab_id of the selected FDH.

        arcpy.AddMessage(f"► Searching for FDH: {cab_id}\n"
                         f"\n")

        if not cab_id:
            arcpy.AddError("❌ No FDH name entered. Please enter a valid cab_id.")
            return None, None, None, None, None, None

        # Retrieve the FDH_Boundary layer from portal
        fdh_item = gis.content.get(fdh_boundary_id)
        if not fdh_item:
            arcpy.AddError("FDH_Boundary layer not found in Portal.")
            return None, None, None, None, None, None

        fdh_layer = fdh_item.layers[0]  # Assuming first layer is correct
        # arcpy.AddMessage(f"Found FDH_Boundary layer: {fdh_layer.url}")

        # Query the layer for the specified cab_id
        query_result = fdh_layer.query(
            where=f"cab_id = '{cab_id}'",
            out_fields="OBJECTID, cab_id, Serv_Area, City_Code, Const_Ven",
            return_geometry=True
        )

        if not query_result.features:
            arcpy.AddError(f"No FDH Boundary found for cab_id: {cab_id}")
            return None, None, None, None, None, None

        # Extract feature details
        selected_feature = query_result.features[0]
        object_id = selected_feature.attributes.get("OBJECTID", "Unknown")
        fdh_geometry = selected_feature.geometry  # Geometry
        serv_area = selected_feature.attributes.get("Serv_Area", "Unknown")
        city_code = selected_feature.attributes.get("City_Code", "Unknown")
        const_ven = selected_feature.attributes.get("Const_Ven", "Unknown")

        # arcpy.AddMessage(f"Found FDH Boundary - OBJECTID: {object_id}")
        # arcpy.AddMessage(f"✅ Retrieved Fields: Serv_Area={serv_area}, City_Code={city_code}, Const_Ven={const_ven}")
        # DEBUG MESSAGE for geometry
        # arcpy.AddMessage(f"✅ Retrieved Geometry: {fdh_geometry} and it's type is {type(fdh_geometry)}")

        return object_id, fdh_geometry, cab_id, serv_area, city_code, const_ven

    except Exception as e:
        arcpy.AddError(f"❌ Error retrieving FDH boundary: {e}")
        return None, None, None, None, None, None


def query_conduit_from_portal(conduit_id, fdh_geometry):
    """Queries a Portal feature layer using its ID, retrieving only features within the selected FDH boundary."""
    try:
        # Retrieve the layer from ArcGIS Portal
        conduit_layer_item = gis.content.get(conduit_id)
        if not conduit_layer_item:
            arcpy.AddError(f"❌ Layer with ID '{conduit_id}' not found in ArcGIS Portal.")
            return 0, 0, 0, 0, 0, 0

        portal_layer = conduit_layer_item.layers[0]  # Assuming the correct layer
        # arcpy.AddMessage(f"✅ Found layer in Portal: {portal_layer.url}")

        # ✅ Ensure geometry is in Esri JSON format
        if isinstance(fdh_geometry, dict):
            geometry_json = json.dumps(fdh_geometry)  # Convert dict to JSON string
        elif hasattr(fdh_geometry, "JSON"):
            geometry_json = fdh_geometry.JSON  # Convert ArcPy Geometry to JSON
        else:
            arcpy.AddError("❌ Invalid geometry format.")
            return 0, 0, 0, 0, 0, 0

        # ✅ Ensure spatial reference matches the Portal layer
        spatial_ref = {"wkid": 4326}  # Default to WGS 1984 (Lat/Lon)
        if "spatialReference" not in fdh_geometry:
            fdh_geometry["spatialReference"] = spatial_ref

        query_result = portal_layer.query(
            geometry_filter=arcgis.geometry.filters.intersects(fdh_geometry),  # Use selected boundary
            geometry_type="esriGeometryPolygon",
            spatial_rel="esriSpatialRelContains",  # Ensures only features within the polygon are selected
            out_fields="UG1FT, LaborFootage, BOMCalc, reareasment, Cond_Diam",
            return_geometry=True
        )

        if not query_result.features:
            arcpy.AddMessage(f"⚠ No features found in {portal_layer.properties.name} within the selected boundary.")
            return 0, 0, 0, 0, 0, 0

        # arcpy.AddMessage(f"✅ Retrieved {len(query_result.features)} features from {portal_layer.properties.name}")

        # Initialize the values for ug1, ug2, total_conduit, and special crossing
        total_ug1ft = 0
        total_ug2ft = 0
        total_1in_conduit = 0
        total_2in_conduit = 0
        total_4in_conduit = 0
        total_ug1ft_reareasment_Y = 0

        for feature in query_result.features:
            properties = feature.attributes  # Extract feature attributes

            ug1ft = properties.get("UG1FT", 0) or 0
            laborfootage = properties.get("LaborFootage", 0) or 0
            bomcalc = properties.get("BOMCalc", 0) or 0
            reareasment = str(properties.get("reareasment", "")).strip().upper()
            cond_diam = str(properties.get("Cond_Diam", "")).strip()

            if "1.25" in cond_diam:
                total_1in_conduit += bomcalc
            elif "2" in cond_diam:
                total_2in_conduit += bomcalc
            elif "4" in cond_diam:
                total_4in_conduit += ug1ft  # Sum UG1 footage for 4" conduit which is only used for special crossings
                arcpy.AddWarning(f'There is a total of {total_4in_conduit: .2f} 4" Conduit Feet!\n '
                                 f'This might indicate a Special Crossing\n'
                                 f'Verify this is accurate and check the Construction Drawing if needed.')

            # Sum total UG1, UG2 and Reareasment
            total_ug1ft += ug1ft
            total_ug2ft += laborfootage

            if reareasment == "Y":
                total_ug1ft_reareasment_Y += ug1ft  # Sum UG1FT where reareasment is 'Y'

        arcpy.AddMessage(f"*** Labor and Conduit Footage Within {cab_id}: ***\n"
                         f"------------------------------------------------------\n"
                         f"► Total UG1 Footage: {total_ug1ft:.2f} feet\n"
                         f"► Total UG2 Footage: {total_ug2ft:.2f} feet\n"
                         f'► Total 1.25" Conduit Footage: {total_1in_conduit:.2f} feet\n'
                         f'► Total 2" Conduit Footage: {total_2in_conduit: .2f} feet\n'
                         f'► Total 4" Conduit Footage: {total_4in_conduit: .2f} feet\n'
                         f"► Total UG1 Footage (Reareasment = 'Y'): {total_ug1ft_reareasment_Y:.2f} feet\n"
                         f'► Total Special Crossing Footage: {total_4in_conduit:.2f} feet\n'
                         f'\n')

        return (total_ug1ft,
                total_ug2ft,
                total_1in_conduit,
                total_ug1ft_reareasment_Y,
                total_4in_conduit,
                total_2in_conduit)

    except Exception as e:
        arcpy.AddError(f"Error: {e}")
        return 0, 0, 0, 0, 0, 0


def query_structures_from_portal(structures_id, fdh_geometry):
    try:
        # Retrieve the layer from ArcGIS Portal
        layer_item = gis.content.get(structures_id)
        if not layer_item:
            arcpy.AddError(f"❌ Layer with ID '{structures_id}' not found in ArcGIS Portal.")
            return 0, 0, 0, 0, 0, 0, 0

        portal_layer = layer_item.layers[0]  # Assuming the correct layer
        # arcpy.AddMessage(f"✅ Found layer in Portal: {portal_layer.url}")

        # ✅ Ensure geometry is in Esri JSON format
        if isinstance(fdh_geometry, dict):
            geometry_json = json.dumps(fdh_geometry)  # Convert dict to JSON string
        elif hasattr(fdh_geometry, "JSON"):
            geometry_json = fdh_geometry.JSON  # Convert ArcPy Geometry to JSON
        else:
            arcpy.AddError("❌ Invalid geometry format.")
            return 0, 0, 0, 0, 0, 0, 0

        # ✅ Ensure spatial reference matches the Portal layer
        spatial_ref = {"wkid": 4326}  # Default to WGS 1984 (Lat/Lon)
        if "spatialReference" not in fdh_geometry:
            fdh_geometry["spatialReference"] = spatial_ref

        query_extent = fdh_geometry
        query_filter = arcgis.geometry.filters.contains(query_extent, sr=102100)

        query_result = portal_layer.query(
            geometry_filter=query_filter,
            out_fields="structuretype", as_df=False)

        if not query_result.features:
            arcpy.AddMessage(f"⚠ No features found in {portal_layer.properties.name} within the selected boundary.")
            return 0, 0, 0, 0, 0, 0, 0

        # arcpy.AddMessage(f"✅ Retrieved {len(query_result.features)} features from {portal_layer.properties.name}")

        structure_counts = defaultdict(int)  # Dictionary to count occurrences of each structure type
        predefined_types = {"FP", "SV", "MV", "LV", "XL", "XSV", "NID Box", "XXL"}  # Predefined structure types to track

        for feature in query_result.features:
            properties = feature.attributes
            structure_type = properties.get("structuretype", "Unknown")
            if structure_type in predefined_types:
                structure_counts[structure_type] += 1

        # **Ensure all structure types are included, even if they have no data**
        fp_count = structure_counts.get("FP", 0)
        sv_count = structure_counts.get("SV", 0)
        mv_count = structure_counts.get("MV", 0)
        lv_count = structure_counts.get("LV", 0)
        xl_count = structure_counts.get("XL", 0)
        xsv_count = structure_counts.get("XSV", 0)
        nid_count = structure_counts.get("NID Box", 0)
        axl_count = structure_counts.get("XXL", 0)

        # Print individual structure type counts
        arcpy.AddMessage(f"*** Structure Counts Within {cab_id} **:\n"
                         f"-----------------------------------------\n"
                         f"► Flowerpots: {fp_count}\n"
                         f"► Small Vaults: {sv_count}\n"
                         f"► Medium Vaults: {mv_count}\n"
                         f"► Large Vaults: {lv_count}\n"
                         f"► XL Vaults: {xl_count}\n"
                         f"► XSV Vaults: {xsv_count}\n"
                         f"► NID Box: {nid_count}\n"
                         f"► AXL: {axl_count}\n"
                         f"\n")

        return (fp_count,
                sv_count,
                mv_count,
                lv_count,
                xl_count,
                xsv_count,
                nid_count,
                axl_count
                )

    except Exception as e:
        arcpy.AddError(f"❌ Error: {e}")
    return 0, 0, 0, 0, 0, 0, 0  # Ensure function always returns all values


def query_splice_sizes_from_portal(splice_enclosure_id, fdh_geometry):
    try:
        # Retrieve the layer from ArcGIS Portal
        layer_item = gis.content.get(splice_enclosure_id)
        if not layer_item:
            arcpy.AddError(f"❌ Layer with ID '{splice_enclosure_id}' not found in ArcGIS Portal.")
            return 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0

        portal_layer = layer_item.layers[0]  # Assuming the correct layer
        # arcpy.AddMessage(f"✅ Found layer in Portal: {portal_layer.url}")

        # ✅ Ensure geometry is in Esri JSON format
        if isinstance(fdh_geometry, dict):
            geometry_json = json.dumps(fdh_geometry)  # Convert dict to JSON string
        elif hasattr(fdh_geometry, "JSON"):
            geometry_json = fdh_geometry.JSON  # Convert ArcPy Geometry to JSON
        else:
            arcpy.AddError("❌ Invalid geometry format.")
            return 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0

        # ✅ Ensure spatial reference matches the Portal layer
        spatial_ref = {"wkid": 4326}  # Default to WGS 1984 (Lat/Lon)
        if "spatialReference" not in fdh_geometry:
            fdh_geometry["spatialReference"] = spatial_ref

        # 🔹 Debugging messages
        # arcpy.AddMessage(f"🔍 Using Spatial Query with Geometry: {geometry_json}")

        query_extent = fdh_geometry
        query_filter = arcgis.geometry.filters.contains(query_extent, sr=102100)

        query_result = portal_layer.query(
            geometry_filter=query_filter,
            out_fields="splicesize, placementtype", as_df=False)

        if not query_result.features:
            arcpy.AddMessage(f"⚠ No features found in {portal_layer.properties.name} within the selected boundary.")
            return 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0

        # arcpy.AddMessage(f"✅ Retrieved {len(query_result.features)} features from {portal_layer.properties.name}")

        # Dictionary to count occurrences of each splice size type
        splicesize_counts = defaultdict(int)

        # Predefined splice sizes to track
        predefined_types = {'Coyote One', '6.5"x17"', '6.5"x22"', '9.5"x28"', '9.5"x19"', "RUNT"}

        # **Initialize hanger_bracket and offset_bracket counts**
        hanger_bracket = 0
        offset_bracket = 0

        # **Iterate through the retrieved features from REST API**
        for feature in query_result.features:
            properties = feature.attributes

            splicesize = properties.get("splicesize", "Unknown")
            placementtype = properties.get("placementtype", "Unknown")

            # Validate splicesize against predefined types
            if splicesize not in predefined_types:
                splicesize = "Unknown"

            # Count occurrences of each splice size
            splicesize_counts[splicesize] += 1

            # Count hanger_bracket where placementtype = AE and splicesize = "Coyote One"
            if placementtype == "AE" and splicesize == "Coyote One":
                hanger_bracket += 1

            # Count offset_bracket where placementtype = AE and splicesize != "Coyote One"
            if placementtype == "AE" and splicesize != "Coyote One":
                offset_bracket += 1

        # **Ensure all splice sizes are included, even if they have no data**
        coyote_count = splicesize_counts.get('Coyote One', 0)
        x17_count = splicesize_counts.get('6.5"x17"', 0)
        x22_count = splicesize_counts.get('6.5"x22"', 0)
        x28_count = splicesize_counts.get('9.5"x28"', 0)
        x19_count = splicesize_counts.get('9.5"x19"', 0)
        runt_count = splicesize_counts.get('RUNT', 0)

        total_closure_count = coyote_count + x17_count + x22_count + x28_count + x19_count + runt_count

        ug_closure_count = abs((hanger_bracket + offset_bracket) - total_closure_count)

        lash_closure_count = hanger_bracket + offset_bracket

        # **Print individual structure type counts**
        arcpy.AddMessage(f"*** Splice Size Counts Within {cab_id}: ***\n"
                         f"--------------------------------------------\n"
                         f"► Coyote One: {coyote_count}\n"
                         f'► 6.5"x17": {x17_count}\n'
                         f'► 6.5"x22": {x22_count}\n'
                         f'► 9.5"x19": {x19_count}\n'
                         f'► 9.5"x28": {x28_count}\n'
                         f'► Runt Closure: {runt_count}\n'
                         f'► Hanger Bracket Count: {hanger_bracket}\n'
                         f'► Offset Bracket Count: {offset_bracket}\n'
                         f'\n')

        return (coyote_count, x17_count, x22_count, x28_count, x19_count, runt_count, total_closure_count,
                hanger_bracket, offset_bracket, ug_closure_count, lash_closure_count)

    except Exception as e:
        arcpy.AddError(f"❌ Error: {e}")
        return 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0


def query_cables_from_portal(cable_id, fdh_geometry):
    try:
        # Retrieve the layer from ArcGIS Portal
        layer_item = gis.content.get(cable_id)
        if not layer_item:
            arcpy.AddError(f"❌ Layer with ID '{cable_id}' not found in ArcGIS Portal.")
            return (0,) * 26

        portal_layer = layer_item.layers[0]  # Assuming the correct layer
        # arcpy.AddMessage(f"✅ Found layer in Portal: {portal_layer.url}")

        # ✅ Ensure geometry is in Esri JSON format
        if isinstance(fdh_geometry, dict):
            geometry_json = json.dumps(fdh_geometry)  # Convert dict to JSON string
        elif hasattr(fdh_geometry, "JSON"):
            geometry_json = fdh_geometry.JSON  # Convert ArcPy Geometry to JSON
        else:
            arcpy.AddError("❌ Invalid geometry format.")
            return (0,) * 26

        # ✅ Ensure spatial reference matches the Portal layer
        spatial_ref = {"wkid": 4326}  # Default to WGS 1984 (Lat/Lon)
        if "spatialReference" not in fdh_geometry:
            fdh_geometry["spatialReference"] = spatial_ref

        query_extent = fdh_geometry
        query_filter = arcgis.geometry.filters.contains(query_extent, sr=102100)

        query_result = portal_layer.query(
            geometry_filter=query_filter,
            out_fields="cable_name, placementtype, fibercount, hierarchy, LengthFT, SpliceSlack, "
                       "SP1, SP2, SP3", as_df=False)

        if not query_result.features:
            arcpy.AddMessage(f"⚠ No features found in {portal_layer.properties.name} within the selected boundary.")
            return (0,) * 26

        # arcpy.AddMessage(f"✅ Retrieved {len(query_result.features)} features from {portal_layer.properties.name}")

        # Dictionary to store summed values for each predefined fiber count and placement type
        fiber_slack_sums = defaultdict(lambda: {"UG": 0.0, "AE": 0.0})
        hierarchy_sums = {"F1_UG": 0.0, "F1_AE": 0.0, "F2_UG": 0.0, "F2_AE": 0.0}
        total_fiber_footage_ug_linear = 0.0  # variable for UG fiber length
        total_fiber_footage_ae_linear = 0.0  # variable for AE fiber length
        unique_cables = {}  # Dictionary to store the first occurrence of each unique cable name
        total_sp3_excluding_f1 = 0  # Initialize SP3 sum excluding 'F1' cables

        # Iterate through the features returned from the portal query
        for feature in query_result.features:
            properties = feature.attributes

            fiber_count = str(properties.get("fibercount", "Unknown"))
            splice_slack = properties.get("SpliceSlack", 0) or 0
            placement_type = properties.get("placementtype", "").strip().upper()
            hierarchy_type = properties.get("hierarchy", "").strip().upper()
            length_ft = properties.get("LengthFT", 0) or 0
            cable_name = properties.get("cable_name", "Unknown")

            # Adding the fiber count to the cable name
            unique_cable_key = f"{cable_name}-{fiber_count}"

            # Sum by unique fiber count and placement type
            if fiber_count and placement_type in ["UG", "AE"]:
                fiber_slack_sums[fiber_count][placement_type] += splice_slack

            # Sum by hierarchy (F1/F2) and placement type
            if hierarchy_type in ["F1", "F2"] and placement_type in ["UG", "AE"]:
                key = f"{hierarchy_type}_{placement_type}"  # Creates "F1_UG", "F2_AE", etc.
                hierarchy_sums[key] += splice_slack

            # Sum total fiber length separately for UG and AE
            if placement_type == "UG":
                total_fiber_footage_ug_linear += length_ft
            elif placement_type == "AE":
                total_fiber_footage_ae_linear += length_ft

            # Store only unique cable names with appended fibercount
            if unique_cable_key not in unique_cables:
                unique_cables[unique_cable_key] = {
                    "SP1": properties.get("SP1", 0) or 0,
                    "SP2": properties.get("SP2", 0) or 0,
                    "SP3": properties.get("SP3", 0) or 0,
                    "hierarchy": hierarchy_type
                }

        # Summing up the values from the unique cables dictionary
        total_sp1 = sum(cable["SP1"] for cable in unique_cables.values())
        total_sp2 = sum(cable["SP2"] for cable in unique_cables.values())

        # Sum SP3 **only for non-F1 cables**, then add a base amount of 24
        for cable in unique_cables.values():
            if cable["hierarchy"] != "F1":
                total_sp3_excluding_f1 += cable["SP3"]

        total_sp3 = total_sp3_excluding_f1 + 24  # ✅ Add base amount of 24
        total_heatshrink = total_sp1 + total_sp2 + total_sp3_excluding_f1

        # Log the list of unique cable names
        unique_cable_list = ", ".join(unique_cables.keys()) if unique_cables else "None"
        arcpy.AddMessage(f"*** Unique Cable Names in {cab_id}: ***\n"
                         f"-------------------------------------------\n"
                         f"►\n"
                         f"{unique_cable_list}")

        # Display results in ArcGIS Pro
        arcpy.AddMessage(f"*** Total Summed Splice Values (Unique Names Only) in {cab_id}: ***\n"
                         f"---------------------------------------------------------------------\n"
                         f"► SPL-1 Total: {total_sp1}\n"
                         f"► SPL-2 Total: {total_sp2}\n"
                         f"► SPL-3 Total (Excluding F1 + 24 base): {total_sp3}\n"
                         f"\n")

        # **Ensure all fiber counts are included, even if they have no data**
        fiber_12_ug = round(fiber_slack_sums["12"]["UG"], 2)
        fiber_12_ae = round(fiber_slack_sums["12"]["AE"], 2)
        fiber_24_ug = round(fiber_slack_sums["24"]["UG"], 2)
        fiber_24_ae = round(fiber_slack_sums["24"]["AE"], 2)
        fiber_48_ug = round(fiber_slack_sums["48"]["UG"], 2)
        fiber_48_ae = round(fiber_slack_sums["48"]["AE"], 2)
        fiber_96_ug = round(fiber_slack_sums["96"]["UG"], 2)
        fiber_96_ae = round(fiber_slack_sums["96"]["AE"], 2)
        fiber_144_ug = round(fiber_slack_sums["144"]["UG"], 2)
        fiber_144_ae = round(fiber_slack_sums["144"]["AE"], 2)
        fiber_288_ug = round(fiber_slack_sums["288"]["UG"], 2)
        fiber_288_ae = round(fiber_slack_sums["288"]["AE"], 2)
        fiber_432_ae = round(fiber_slack_sums["432"]["AE"], 2)
        fiber_432_ug = round(fiber_slack_sums["432"]["UG"], 2)

        # Calculate total fiber footage (sum of all SpliceSlack values)
        total_fiber_footage_ug = sum(
            [fiber_12_ug, fiber_24_ug, fiber_48_ug, fiber_96_ug, fiber_144_ug, fiber_288_ug, fiber_432_ug])
        total_fiber_footage_ae = sum(
            [fiber_12_ae, fiber_24_ae, fiber_48_ae, fiber_96_ae, fiber_144_ae, fiber_288_ae, fiber_432_ae])

        # **Extract F1/F2 values from hierarchy sums**
        total_f1_ug = round(hierarchy_sums["F1_UG"], 2)
        total_f1_ae = round(hierarchy_sums["F1_AE"], 2)
        total_f2_ug = round(hierarchy_sums["F2_UG"], 2)
        total_f2_ae = round(hierarchy_sums["F2_AE"], 2)

        # Round total fiber length
        total_fiber_footage_ug_linear = round(total_fiber_footage_ug_linear, 2)
        total_fiber_footage_ae_linear = round(total_fiber_footage_ae_linear, 2)

        # Output Messages
        arcpy.AddMessage(f"*** Fiber Footage by Count Within {cab_id}: ***\n"
                         f"------------------------------------------------\n"
                         f"► Fiber 12 UG: {fiber_12_ug} ft | AE: {fiber_12_ae} ft\n"
                         f"► Fiber 24 UG: {fiber_24_ug} ft | AE: {fiber_24_ae} ft\n"
                         f"► Fiber 48 UG: {fiber_48_ug} ft | AE: {fiber_48_ae} ft\n"
                         f"► Fiber 96 UG: {fiber_96_ug} ft | AE: {fiber_96_ae} ft\n"
                         f"► Fiber 144 UG: {fiber_144_ug} ft | AE: {fiber_144_ae} ft\n"
                         f"► Fiber 288 UG: {fiber_288_ug} ft | AE: {fiber_288_ae} ft\n"
                         f"► Fiber 432 UG: {fiber_432_ug} ft | AE: {fiber_432_ae} ft\n"
                         f"\n")

        return (
            fiber_12_ug, fiber_12_ae, fiber_24_ug, fiber_24_ae,
            fiber_48_ug, fiber_48_ae, fiber_96_ug, fiber_96_ae,
            fiber_144_ug, fiber_144_ae, fiber_288_ug, fiber_288_ae,
            total_fiber_footage_ug, total_fiber_footage_ae,
            total_f1_ug, total_f1_ae, total_f2_ug, total_f2_ae,
            total_fiber_footage_ug_linear, total_fiber_footage_ae_linear,
            total_sp1, total_sp2, total_sp3_excluding_f1, total_heatshrink, fiber_432_ug, fiber_432_ae)

    except Exception as e:
        arcpy.AddError(f"❌ Error: {e}")
        return (0,) * 26


def query_slackloops_from_portal(slackloop_id, fdh_geometry):
    try:
        # Retrieve the layer from ArcGIS Portal
        layer_item = gis.content.get(slackloop_id)
        if not layer_item:
            arcpy.AddError(f"❌ Layer with ID '{slackloop_id}' not found in ArcGIS Portal.")
            return {}, 0, 0

        portal_layer = layer_item.layers[0]  # Assuming the correct layer
        # arcpy.AddMessage(f"✅ Found layer in Portal: {portal_layer.url}")

        # ✅ Ensure geometry is in Esri JSON format
        if isinstance(fdh_geometry, dict):
            geometry_json = json.dumps(fdh_geometry)  # Convert dict to JSON string
        elif hasattr(fdh_geometry, "JSON"):
            geometry_json = fdh_geometry.JSON  # Convert ArcPy Geometry to JSON
        else:
            arcpy.AddError("❌ Invalid geometry format.")
            return {}, 0, 0

        # ✅ Ensure spatial reference matches the Portal layer
        spatial_ref = {"wkid": 4326}  # Default to WGS 1984 (Lat/Lon)
        if "spatialReference" not in fdh_geometry:
            fdh_geometry["spatialReference"] = spatial_ref

        query_extent = fdh_geometry
        query_filter = arcgis.geometry.filters.contains(query_extent, sr=102100)

        query_result = portal_layer.query(
            geometry_filter=query_filter,
            out_fields="cable_capacity, placement, loop_length, type", as_df=False)

        if not query_result.features:
            arcpy.AddMessage(f"⚠ No features found in {portal_layer.properties.name} within the selected boundary.\n")
            return {}, 0, 0

        # arcpy.AddMessage(f"✅ Retrieved {len(query_result.features)} features from {portal_layer.properties.name}")

        # Initialize dictionaries for storing summed values
        slackloop_sums = defaultdict(lambda: {"UG": 0, "AE": 0, "Total": 0})
        total_ug_slackloops = 0  # ✅ Track total UG slackloops separately
        total_ae_slackloops = 0  # ✅ Track total AE slackloops separately

        # **Iterate through the retrieved features from REST API**
        for feature in query_result.features:
            properties = feature.attributes

            cable_capacity = str(properties.get("cable_capacity") or "Unknown")
            placement = str(properties.get("placement", "Unknown")).upper()
            loop_length = (properties.get("loop_length", 0) or 0)
            loop_type = str(properties.get("type") or "").strip().upper()
            if not loop_type:
                loop_type = "UNKNOWN"

            # ✅ Filter for 'Maintenance Loop' only
            if loop_type != "MAINTENANCE LOOP":
                continue

            # Sum values for UG and AE separately
            if placement == "UG":
                slackloop_sums[cable_capacity]["UG"] += loop_length
                total_ug_slackloops += 1  # ✅ Count UG slackloops separately
            elif placement == "AE":
                slackloop_sums[cable_capacity]["AE"] += loop_length
                total_ae_slackloops += 1  # ✅ Count AE slackloops separately

            # Sum total length (UG + AE)
            slackloop_sums[cable_capacity]["Total"] = (
                    slackloop_sums[cable_capacity]["UG"] + slackloop_sums[cable_capacity]["AE"]
            )  # ✅ Explicitly sum UG + AE to ensure accuracy

        arcpy.AddMessage(f"*** Slackloop Sums by Fiber Count and Placement Within {cab_id}: ***")
        for cap, values in slackloop_sums.items():
            arcpy.AddMessage(
                f"►  - {cap}: UG = {values['UG']} ft, AE = {values['AE']} ft, Total = {values['Total']} ft\n")
            arcpy.AddMessage("\n")
            arcpy.AddMessage("\n")

        return slackloop_sums, total_ug_slackloops, total_ae_slackloops

    except Exception as e:
        arcpy.AddError(f"❌ Error processing Slackloop features: {e}")
        return {}, 0, 0  # Ensure function always returns expected values


def query_strand_and_poles_from_portal(strand_id, poles_id, conduit_id, fdh_geometry):
    try:
        # Retrieve the strand layer from ArcGIS Portal
        strand_layer_item = gis.content.get(strand_id)
        if not strand_layer_item:
            arcpy.AddError(f"❌ Layer with ID '{strand_id}' not found in ArcGIS Portal.")
            return 0, 0, 0, 0, 0

        # Retrieve the pole layer from ArcGIS Portal
        pole_layer_item = gis.content.get(poles_id)
        if not pole_layer_item:
            arcpy.AddError(f"❌ Layer with ID '{poles_id}' not found in ArcGIS Portal.")
            return 0, 0, 0, 0, 0

        # Retrieve the layer from ArcGIS Portal
        conduit_layer_item = gis.content.get(conduit_id)
        if not conduit_layer_item:
            arcpy.AddError(f"❌ Layer with ID '{conduit_id}' not found in ArcGIS Portal.")
            return 0, 0, 0, 0, 0

        portal_strand_layer = strand_layer_item.layers[0]  # Assuming the correct layer index
        portal_pole_layer = pole_layer_item.layers[0]  # Assuming the correct layer index
        portal_conduit_layer = conduit_layer_item.layers[0]

        # ✅ Ensure geometry is in Esri JSON format
        if isinstance(fdh_geometry, dict):
            geometry_json = json.dumps(fdh_geometry)  # Convert dict to JSON string
        elif hasattr(fdh_geometry, "JSON"):
            geometry_json = fdh_geometry.JSON  # Convert ArcPy Geometry to JSON
        else:
            arcpy.AddError("❌ Invalid geometry format.")
            return {}

        # ✅ Ensure spatial reference matches the Portal layer
        spatial_ref = {"wkid": 4326}  # Default to WGS 1984 (Lat/Lon)
        if "spatialReference" not in fdh_geometry:
            fdh_geometry["spatialReference"] = spatial_ref

        # Query the strand layer against the FDH-Boundary geometry
        query_result_strand = portal_strand_layer.query(
            geometry_filter=arcgis.geometry.filters.contains(fdh_geometry, sr=102100),
            return_geometry=True,
            as_df=False
        )

        if not query_result_strand.features:
            arcpy.AddMessage(f"\n ⚠ No features found in {portal_strand_layer.properties.name} "
                             f"within the selected boundary.\n"
                             f"\n")
            return 0, 0, 0, 0, 0

        total_strand_ftg = 0  # Total strand footage
        total_strand_ftg_reareasment_y = 0  # Strand footage where reareasment = 'Y'
        strand_geometries = []  # Store strand geometries for intersection check

        # Iterate through the retrieved strand features from the Portal
        for feature in query_result_strand.features:
            properties = feature.attributes
            strand_geometry = feature.geometry
            strand_ftg = properties.get("calcfootage", 0) or 0
            reareasment = str(properties.get("reareasment", "UNKNOWN")).strip().upper()

            total_strand_ftg += strand_ftg
            if reareasment == "Y":
                total_strand_ftg_reareasment_y += strand_ftg

            if strand_geometry:
                strand_geometries.append(strand_geometry)  # Store for pole intersection check

        intersecting_poles = []  # Store pole features that intersect strands

        for strand_geom in strand_geometries:
            query_result_poles = portal_pole_layer.query(
                geometry_filter=arcgis.geometry.filters.intersects(strand_geom, sr=102100),
                out_fields="MR_Level",
                return_geometry=True,
                as_df=False
            )
            intersecting_poles.extend(query_result_poles.features)  # Append results

        total_pole_count = len(intersecting_poles)  # Total poles intersecting strands

        mr_filtered_pole_count = 0

        for pole in intersecting_poles:
            mr_level = pole.attributes.get("MR_Level", None)  # Extract MR_Level
            if mr_level in [1, 2]:  # Check if MR_Level is 1 or 2
                mr_filtered_pole_count += 1

        total_strand_ftg *= 1.10
        total_strand_ftg_reareasment_y *= 1.10

        # Get pole features within the FDH Boundary
        pole_features = portal_pole_layer.query(
            geometry_filter=arcgis.geometry.filters.contains(fdh_geometry, sr=102100),
            out_fields="OBJECTID",  # Only field needed since we just want a count of poles
            return_geometry=True,  # required for intersect
            as_df=False
        ).features

        uguard_adapter = 0

        # Iterating through the retrieved pole features
        for pole in pole_features:
            pole_geom = pole.geometry
            conduits_at_pole = portal_conduit_layer.query(  # Get the ducts intersecting poles
                geometry_filter=arcgis.geometry.filters.intersects(pole_geom, sr=102100),
                out_fields="duct_count",
                return_geometry=False
            ).features

            # Initialize variable to count the ducts at the poles
            duct_sum = 0

            # Iterate through the conduits at the poles and sum the ducts
            for conduit in conduits_at_pole:
                duct_cnt = conduit.attributes.get("duct_count", 0) or 0
                try:
                    duct_sum += int(duct_cnt)
                except:
                    pass  # skip if not convertible

            if duct_sum >= 3:
                uguard_adapter += 1  # increment the uguard for each pole with 3 or more ducts

        arcpy.AddMessage(f"\n*** Strand and Poles Within {cab_id}: ***\n"
                         f"--------------------------------------------\n"
                         f"► Total Strand Footage (including sag): {total_strand_ftg:.2f} feet\n"
                         f"► Total Strand Footage (where reareasment='Y', including sag): "
                         f"{total_strand_ftg_reareasment_y:.2f} feet\n"
                         f"► Total Poles Intersecting Strand: {total_pole_count}\n"
                         f"► Total Poles Requiring Make Ready: {mr_filtered_pole_count}\n"
                         f"\n")

        return (total_strand_ftg,
                total_strand_ftg_reareasment_y,
                total_pole_count,
                mr_filtered_pole_count,
                uguard_adapter
                )

    except Exception as e:
        arcpy.AddError(f"❌ Error processing Strand and Pole features: {e}")
        return 0, 0, 0, 0, 0  # Ensure function always returns three values


def query_cabinets_from_portal(passive_id, active_id, fdh_geometry):
    try:
        # Retrieve the passive_cabinet layer from ArcGIS Portal
        passive_layer_item = gis.content.get(passive_id)
        if not passive_layer_item:
            arcpy.AddError(f"❌ Layer with ID '{passive_id}' not found in ArcGIS Portal.")
            return 0, 0, 0, 0, 0

        # Retrieve the active_cabinet layer from ArcGIS Portal
        active_layer_item = gis.content.get(active_id)
        if not active_layer_item:
            arcpy.AddError(f"❌ Layer with ID '{active_id}' not found in ArcGIS Portal.")
            return 0, 0, 0, 0, 0

        portal_passive_layer = passive_layer_item.layers[0]  # Assuming the correct layer index
        portal_active_layer = active_layer_item.layers[0]  # Assuming the correct layer index
        # arcpy.AddMessage(f"✅ Found layer in Portal: {portal_passive_layer.url}")
        # arcpy.AddMessage(f"✅ Found layer in Portal: {portal_active_layer.url}")

        # # ✅ Ensure geometry is in Esri JSON format
        # if isinstance(fdh_geometry, dict):
        #     geometry_json = json.dumps(fdh_geometry)  # Convert dict to JSON string
        # elif hasattr(fdh_geometry, "JSON"):
        #     geometry_json = fdh_geometry.JSON  # Convert ArcPy Geometry to JSON
        # else:
        #     arcpy.AddError("❌ Invalid geometry format.")
        #     return {}

        # ✅ Ensure spatial reference matches the Portal layer
        spatial_ref = {"wkid": 4326}  # Default to WGS 1984 (Lat/Lon)
        if "spatialReference" not in fdh_geometry:
            fdh_geometry["spatialReference"] = spatial_ref

        # Query the strand layer against the FDH-Boundary geometry
        query_result_passive = portal_passive_layer.query(
            geometry_filter=arcgis.geometry.filters.contains(fdh_geometry, sr=102100),
            out_fields="Cab_Size",
            return_geometry=True,
            as_df=False
        )

        if not query_result_passive.features:
            arcpy.AddMessage(f"⚠ No features found in {portal_passive_layer.properties.name} "
                             f"within the selected boundary.")
            return 0, 0, 0, 0, 0

        # arcpy.AddMessage(f"✅ Retrieved {len(query_result_passive.features)} "
        # f"features from {portal_passive_layer.properties.name}")

        passive_counts = defaultdict(int)  # Dictionary to count occurrences of each passive size
        predefined_types = {"144", "288", "432", "576"}  # Predefined passive cabinet sizes to track

        for feature in query_result_passive.features:
            properties = feature.attributes
            passive_size = properties.get("Cab_Size", "Unknown")
            if passive_size in predefined_types:
                passive_counts[passive_size] += 1

        passive_144 = passive_counts.get("144", 0)
        passive_288 = passive_counts.get("288", 0)
        passive_432 = passive_counts.get("432", 0)
        passive_576 = passive_counts.get("576", 0)

        # Output Messages
        arcpy.AddMessage(f"*** Passive Cabinets and Sizes Within {cab_id}: ***\n"
                         f"----------------------------------------------\n"
                         f"► 144 Passive Cabinets within FDH_Boundary: {passive_144}\n"
                         f"► 288 Passive Cabinets within FDH_Boundary: {passive_288}\n"
                         f"► 432 Passive Cabinets within FDH_Boundary: {passive_432}\n"
                         f"► 576 Passive Cabinets within FDH_Boundary: {passive_576}\n"
                         f"\n")

        # Query the active_cabinet layer from the Portal
        query_result_active = portal_active_layer.query(
            geometry_filter=arcgis.geometry.filters.contains(fdh_geometry, sr=102100),
            out_fields="*",
            return_geometry=True,
            as_df=False
        )

        if not query_result_active.features:
            arcpy.AddMessage(f"⚠ No features found in {portal_active_layer.properties.name} "
                             f"within the selected boundary.")

            return passive_144, passive_288, passive_432, passive_576, 0

        active_cabinet_count = len(query_result_active.features)

        # Output Messages
        arcpy.AddMessage(f"*** Total Active Cabinets Within {cab_id}: ***\n"
                         f"------------------------------------------------\n"
                         f"► Active Cabinet Count: {active_cabinet_count}\n"
                         f"\n")

        return passive_144, passive_288, passive_432, passive_576, active_cabinet_count

    except Exception as e:
        arcpy.AddError(f"❌ Error processing Active Cabinets or Passive Cabinets: {e}")
        return 0, 0, 0, 0, 0  # Ensure function always returns a value


def query_risers_from_portal(riser_id, fdh_geometry):
    try:
        # Retrieve the riser layer from ArcGIS Portal
        riser_layer_item = gis.content.get(riser_id)
        if not riser_layer_item:
            arcpy.AddError(f"❌ Layer with ID '{riser_id}' not found in ArcGIS Portal.")
            return 0

        portal_riser_layer = riser_layer_item.layers[0]  # Assuming the correct layer index
        # arcpy.AddMessage(f"✅ Found layer in Portal: {portal_riser_layer.url}")


        # ✅ Ensure spatial reference matches the Portal layer
        spatial_ref = {"wkid": 4326}  # Default to WGS 1984 (Lat/Lon)
        if "spatialReference" not in fdh_geometry:
            fdh_geometry["spatialReference"] = spatial_ref

        # Query the riser layer against the FDH-Boundary geometry
        query_result_riser = portal_riser_layer.query(
            geometry_filter=arcgis.geometry.filters.contains(fdh_geometry, sr=102100),
            out_fields="*",
            return_geometry=True,
            as_df=False
        )

        if not query_result_riser.features:
            arcpy.AddMessage(f"⚠ No features found in {portal_riser_layer.properties.name} "
                             f"within the selected boundary.\n"
                             f"\n")
            return 0

        total_risers = len(query_result_riser.features)

        arcpy.AddMessage(f"*** Total Risers Within {cab_id}: ***\n"
                         f"----------------------------------------\n"
                         f"► Riser Count: {total_risers}\n"
                         f"\n")

        return total_risers

    except Exception as e:
        arcpy.AddError(f"❌ Error processing Active Cabinets or Passive Cabinets: {e}")
        return 0  # Ensure function always returns a value


def query_guys_from_portal(guys_id, fdh_geometry):
    try:
        # Retrieve the riser layer from ArcGIS Portal
        guys_layer_item = gis.content.get(guys_id)
        if not guys_layer_item:
            arcpy.AddError(f"❌ Layer with ID '{guys_id}' not found in ArcGIS Portal.")
            return 0, 0, 0

        portal_guys_layer = guys_layer_item.layers[0]  # Assuming the correct layer index

        # # ✅ Ensure geometry is in Esri JSON format
        # if isinstance(fdh_geometry, dict):
        #     geometry_json = json.dumps(fdh_geometry)  # Convert dict to JSON string
        # elif hasattr(fdh_geometry, "JSON"):
        #     geometry_json = fdh_geometry.JSON  # Convert ArcPy Geometry to JSON
        # else:
        #     arcpy.AddError("❌ Invalid geometry format.")
        #     return {}

        # ✅ Ensure spatial reference matches the Portal layer
        spatial_ref = {"wkid": 4326}  # Default to WGS 1984 (Lat/Lon)
        if "spatialReference" not in fdh_geometry:
            fdh_geometry["spatialReference"] = spatial_ref

        query_result_guys = portal_guys_layer.query(
            geometry_filter=arcgis.geometry.filters.contains(fdh_geometry, sr=102100),
            out_fields="Guy_Type",
            return_geometry=True,
            as_df=False
        )

        if not query_result_guys.features:
            arcpy.AddMessage(f"⚠ No features found in {portal_guys_layer.properties.name} "
                             f"within the selected boundary.\n"
                             f"\n")
            return 0, 0, 0

        guy_counts = defaultdict(int)  # Dictionary to count occurrences of each structure type
        predefined_types = {"Down", "Dirt", "Rock"}  # Predefined structure types to track

        for feature in query_result_guys.features:
            properties = feature.attributes
            guy_type = properties.get("Guy_Type", "Unknown")
            if guy_type in predefined_types:
                guy_counts[guy_type] += 1

        # **Ensure all structure types are included, even if they have no data**
        down_count = guy_counts.get("Down", 0)
        dirt_count = guy_counts.get("Dirt", 0)
        rock_count = guy_counts.get("Rock", 0)

        # Print individual structure type counts
        arcpy.AddMessage(f"*** Anchor Counts Within {cab_id} **:\n"
                         f"-----------------------------------------\n"
                         f"► Down Guys: {down_count}\n"
                         f"► Dirt Anchors: {dirt_count}\n"
                         f"► Rock Anchors: {rock_count}\n"
                         f"\n")

        return down_count, dirt_count, rock_count

    except Exception as e:
        arcpy.AddError(f"❌ Error: {e}")
    return 0, 0, 0


def query_drops_from_portal(drop_id, fdh_geometry):
    try:
        # Retrieve the drop layer from ArcGIS Portal
        drop_layer_item = gis.content.get(drop_id)
        if not drop_layer_item:
            arcpy.AddError(f"❌ Layer with ID '{drop_id}' not found in ArcGIS Portal.")
            return 0, 0, 0

        portal_drop_layer = drop_layer_item.layers[0]  # Assuming the correct layer index
        # arcpy.AddMessage(f"✅ Found layer in Portal: {portal_drop_layer.url}")

        # ✅ Ensure spatial reference matches the Portal layer
        spatial_ref = {"wkid": 4326}  # Default to WGS 1984 (Lat/Lon)
        if "spatialReference" not in fdh_geometry:
            fdh_geometry["spatialReference"] = spatial_ref

        # Query the drop layer against the FDH-Boundary geometry
        query_result_drops = portal_drop_layer.query(
            geometry_filter=arcgis.geometry.filters.contains(fdh_geometry, sr=102100),
            out_fields="*",
            return_geometry=True,
            as_df=False
        )

        if not query_result_drops.features:
            arcpy.AddMessage(f"⚠ No features found in {portal_drop_layer.properties.name}"
                             f"within the selected boundary.")
            return 0, 0, 0

        # Get the count of retrieved drops
        drop_count = len(query_result_drops.features)

        if drop_count == 0:
            arcpy.AddWarning("⚠ No Drops found within the selected boundary.")
            return 0, 0, 0

        # Initialize counters for drops stats
        count_over_600ft = 0
        total_calcfootage = 0

        # **Iterate through the retrieved features from the Portal**
        for feature in query_result_drops.features:
            properties = feature.attributes
            # drop_geometry = feature.geometry  # Extract geometry

            calcfootage = properties.get("calcfootage", 0) or 0  # Ensure None values default to 0
            total_calcfootage += calcfootage

            if calcfootage > 600:
                count_over_600ft += 1

        # Calculate average calcfootage length
        average_calcfootage = total_calcfootage / drop_count if drop_count > 0 else 0

        # Output Messages
        arcpy.AddMessage(f"*** Drop Stats Within {cab_id}: ***\n"
                         f"-------------------------------------\n"
                         f"► Drops over 600ft: {count_over_600ft}\n"
                         f"► Average Drop Length: {average_calcfootage:.2f} ft\n"
                         f"\n")

        return count_over_600ft, average_calcfootage, drop_count

    except Exception as e:
        arcpy.AddError(f"❌ Error analyzing drops: {e}")
        return 0, 0, 0  # Ensure function always returns expected values


if __name__ == "__main__":

    fdh_boundary_id = "577f024964b844b7836402bf1f84b01f"
    conduit_id = "cd6de7b04ed144fe833317fd7fd7731e"
    structures_id = "47f9081030fa4c50a9ea13b12e5a27e8"
    splice_enclosure_id = "65482deab3594b5d9c572b8b41715519"
    cable_id = "d8380eadf1514800ba303842456798b1"
    slackloop_id = "8124b9d500c240749221ece33c785763"
    strand_id = "a1950b90b7214b30867bd57bb7760626"
    poles_id = "bc21b517ca3b4594b27b41ede3b5eb6a"
    passive_id = "f1bd84729048403fa02153fe1af54bc9"
    active_id = "8a42d8a5d7b649109101b15647a2235d"
    riser_id = "8f42330d5a264cdca3bd692cc4b268fe"
    drop_id = "9f7962eb211a451da43748fd21122911"
    mdu_boundary_id = "54ec733402cc40c3b95415cdf5005a8a"
    do_not_build_id = "1c0e4200a5c84664b8c73ccda21acc08"
    address_master_id = "dfb329f0de874dbca01eee76133c250d"
    guys_id = "3de8975d28034f53a2680d51279bae67"
    addresses_id = "0e3a2268b3434e2a8d39a208eba032a6"

    # Returning attributes from the selected FDH_Boundary
    object_id, fdh_geometry, cab_id, serv_area, city_code, const_ven = (
        fdh_boundary_selection(fdh_boundary_id))

    # Returning calculations from the conduit within the selected FDH_Boundary
    total_ug1ft, total_ug2ft, total_1in_conduit, total_ug1ft_reareasment_Y, total_4in_conduit, total_2in_conduit = (
        query_conduit_from_portal(conduit_id, fdh_geometry))

    fp_count, sv_count, mv_count, lv_count, xl_count, xsv_count, nid_count, axl_count = (
        query_structures_from_portal(structures_id, fdh_geometry))

    (coyote_count, x17_count, x22_count, x28_count, x19_count, runt_count, total_closure_count,
        hanger_bracket, offset_bracket, ug_closure_count, lash_closure_count) = (
            query_splice_sizes_from_portal(splice_enclosure_id, fdh_geometry))

    (fiber_12_ug, fiber_12_ae, fiber_24_ug, fiber_24_ae, fiber_48_ug, fiber_48_ae, fiber_96_ug, fiber_96_ae,
        fiber_144_ug, fiber_144_ae, fiber_288_ug, fiber_288_ae, total_fiber_footage_ug, total_fiber_footage_ae,
        total_f1_ug, total_f1_ae, total_f2_ug, total_f2_ae, total_fiber_footage_ug_linear,
        total_fiber_footage_ae_linear, total_sp1, total_sp2, total_sp3_excluding_f1, total_heatshrink,
     fiber_432_ug, fiber_432_ae) = (
            query_cables_from_portal(cable_id, fdh_geometry))

    slackloop_sums, total_ug_slackloops, total_ae_slackloops = (
        query_slackloops_from_portal(slackloop_id, fdh_geometry))

    total_strand_ftg, total_strand_ftg_reareasment_y, total_pole_count, mr_filtered_pole_count, uguard_adapter = (
        query_strand_and_poles_from_portal(strand_id, poles_id, conduit_id, fdh_geometry))

    down_count, dirt_count, rock_count = query_guys_from_portal(guys_id, fdh_geometry)

    total_anchors = down_count + dirt_count + rock_count

    passive_144, passive_288, passive_432, passive_576, active_cabinet_count = (
        query_cabinets_from_portal(passive_id, active_id, fdh_geometry))

    total_risers = query_risers_from_portal(riser_id, fdh_geometry)

    count_over_600ft, average_calcfootage, drop_count = query_drops_from_portal(drop_id, fdh_geometry)

    total_addresses, total_hhp_mdu, total_dnb_addresses, mdu_boundary_count, dnb_boundary_count = (
        count_addresses(fdh_geometry))

    arcpy.AddMessage(f"*** HHPs Within {cab_id}: ***\n"
                     f"---------------------------------\n"
                     f"► Total HHPs with Drops: {drop_count}\n"
                     f"► Total HHPs within MDU Boundaries within {cab_id}: {total_hhp_mdu}\n"
                     f"► Total HHP's within DNB Boundaries within {cab_id}: {total_dnb_addresses}\n"
                     f"\n")

    # Various calculation for the BOM Template
    est_total_miles = round(((total_strand_ftg + total_ug1ft) / 5280), 2)
    ae_bom_miles = round((total_strand_ftg / 5280), 2)
    ug_bom_miles = round((total_ug1ft / 5280), 2)
    if total_strand_ftg + total_ug1ft == 0:
        percent_ae = 0
        percent_ug = 0
        arcpy.AddWarning("- The formula for calculating AE Percentage is "
                         "  Dividing by Zero!\n"
                         " - Recalculate geometry of Strand and Conduit")
    else:
        percent_ae = (total_strand_ftg / (total_strand_ftg + total_ug1ft)) * 100
        percent_ug = (total_ug1ft / (total_strand_ftg + total_ug1ft)) * 100

    # Calculates total HHP which is F2 fed plus F1 fed
    total_hhp = drop_count + total_hhp_mdu
    # Counts total snowshoes (a pair)
    snowshoes = total_ae_slackloops + lash_closure_count
    # Calculates the additional fiber lashed to strand
    pfa_2 = total_fiber_footage_ae - total_strand_ftg
    if pfa_2 < 0:
        arcpy.AddWarning(f"⚠️ PFA-2 is negative!\n"
                         f"- This may indicate there are still strand features in a 100% UG Boundary\n"
                         f" OR\n"
                         f"- This may indicate the cable and strand footage need to be re-calculated\n"
                         f"- Please review the data in the FDH Boundary and try again.\n"
                         f" ** Setting PFA-2 to 0 **")
        pfa_2 = 0

    # Extract UG values for different fiber capacities for slackloops.
    # This allows the total UG cable footage for the 'pull fiber' amount in the BOM to account for slacks
    slackloop_12_ug = slackloop_sums.get("12", {}).get("UG", 0)
    slackloop_24_ug = slackloop_sums.get("24", {}).get("UG", 0)
    slackloop_48_ug = slackloop_sums.get("48", {}).get("UG", 0)
    slackloop_96_ug = slackloop_sums.get("96", {}).get("UG", 0)
    slackloop_144_ug = slackloop_sums.get("144", {}).get("UG", 0)
    slackloop_288_ug = slackloop_sums.get("288", {}).get("UG", 0)
    slackloop_432_ug = slackloop_sums.get("432", {}).get("UG", 0)

    # Extract Slackloop UG + AE values
    # This allows the summation of all cable footage + maintenance loops.
    # The splice slack is already accounted for.
    slackloop_12 = slackloop_sums.get("12", {}).get("UG", 0) + slackloop_sums.get("12", {}).get("AE", 0)
    slackloop_24 = slackloop_sums.get("24", {}).get("UG", 0) + slackloop_sums.get("24", {}).get("AE", 0)
    slackloop_48 = slackloop_sums.get("48", {}).get("UG", 0) + slackloop_sums.get("48", {}).get("AE", 0)
    slackloop_96 = slackloop_sums.get("96", {}).get("UG", 0) + slackloop_sums.get("96", {}).get("AE", 0)
    slackloop_144 = slackloop_sums.get("144", {}).get("UG", 0) + slackloop_sums.get("144", {}).get("AE", 0)
    slackloop_288 = slackloop_sums.get("288", {}).get("UG", 0) + slackloop_sums.get("288", {}).get("AE", 0)
    slackloop_432 = slackloop_sums.get("432", {}).get("UG", 0) + slackloop_sums.get("432", {}).get("AE", 0)

    # Add Slackloop footage to fiber footage (final total per fiber count regardless of placement)
    # This is for the material section of the BOM Template
    fiber_12 = (slackloop_12 + fiber_12_ug + fiber_12_ae)
    fiber_24 = (slackloop_24 + fiber_24_ug + fiber_24_ae)
    fiber_48 = (slackloop_48 + fiber_48_ug + fiber_48_ae)
    fiber_96 = (slackloop_96 + fiber_96_ug + fiber_96_ae)
    fiber_144 = (slackloop_144 + fiber_144_ug + fiber_144_ae)
    fiber_288 = (slackloop_288 + fiber_288_ug + fiber_288_ae)
    fiber_432 = (slackloop_432 + fiber_432_ug + fiber_432_ae)

    # Step 15: Sum F1 and F2 fiber cable footage
    total_f1 = total_f1_ug + total_f1_ae
    total_f2 = total_f2_ug + total_f2_ae
    total_f1_miles = round((total_f1 / 5280), 2)
    total_f1_ae_miles = round((total_f1_ae / 5280), 2)
    total_f1_ug_miles = round((total_f1_ug / 5280), 2)
    total_f2_miles = round((total_f2 / 5280), 2)
    total_f2_ae_miles = round((total_f2_ae / 5280), 2)
    total_f2_ug_miles = round((total_f2_ug / 5280), 2)
    total_ae_ftg = fiber_12_ae + fiber_24_ae + fiber_48_ae + fiber_96_ae + fiber_144_ae + fiber_288_ae + fiber_432_ae
    total_ug_ftg = fiber_12_ug + fiber_24_ug + fiber_48_ug + fiber_96_ug + fiber_144_ug + fiber_288_ug + fiber_432_ug

    # Add the total UG fiber footage with the UG slackloop footage (Maintenance Loops) to get the full amount of
    # pull-through fiber footage
    pfd_1 = total_fiber_footage_ug + (slackloop_12_ug + slackloop_24_ug + slackloop_48_ug + slackloop_96_ug +
                                      slackloop_144_ug + slackloop_288_ug + slackloop_432_ug)

    # 1.25" Coupler Calculation
    conduit_couplers_1in = round(total_1in_conduit / 300)
    if conduit_couplers_1in < 1:
        conduit_couplers_1in = 0

    # 2" Coupler Calculation
    conduit_couplers_2in = round(total_2in_conduit / 300)
    if conduit_couplers_2in < 1:
        conduit_couplers_2in = 0

    # 4" Coupler Calculation
    conduit_couplers_4in = round(total_4in_conduit / 300)
    if conduit_couplers_4in < 1:
        conduit_couplers_4in = 0

    # Passive_cabinet_count
    total_cabinets = passive_144 + passive_288 + passive_432 + passive_576

    if total_cabinets == 0:
        arcpy.AddWarning(f"⚠️ The total cabinets in the FDH Boundary is zero!\n"
                         f"- This may indicate the Cabinet Size attribute is not populated.\n"
                         f"- Please review the data in the map and try again!")

    # Total linear footage
    total_linear_footage = total_fiber_footage_ug_linear + total_fiber_footage_ae_linear

    # Total Strand Footage plus 25' Per Anchor
    anchor_strand = total_anchors * 25
    total_strand_ftg = total_strand_ftg + anchor_strand  # Reassigns total strand
    # footage to include the anchor strand footage

    # Grounded Poles
    grounded_poles = total_pole_count * 0.25

    # Tree Trimming
    tree_trimming = total_strand_ftg * 0.175

    # Engineering Project Manager - Run Time Engineering
    e_epmrt_1 = 10

    # Lashing Wire Calculation
    lashing_wire = round((total_strand_ftg * 1.5), 2)

    # Special Crossing footage is 4" conduit + 50
    if total_4in_conduit > 0:
        special_crossing = total_4in_conduit + 50
    else:
        special_crossing = 0

    # Warnings
    if uguard_adapter > 0 and total_risers == 0:
        arcpy.AddWarning(f"There are {uguard_adapter} UGuard Adapters but {total_risers} Risers. "
                         f"Verify Risers in FDH Boundary...")

    if total_anchors == 0 and total_strand_ftg > 0:
        arcpy.AddWarning(f"There are Strand Features intersecting Pole Features, but no Anchors! "
                         f"Please verify FDH Boundary Data...")

    # Store all variables in a dictionary to export to the BOM Template
    values_dict = {
            "total_ug1ft": total_ug1ft,
            "total_ug2ft": total_ug2ft,
            "total_1in_conduit": total_1in_conduit,
            "total_2in_conduit": total_2in_conduit,
            "total_4in_conduit": total_4in_conduit,
            "conduit_couplers_1in": conduit_couplers_1in,
            "conduit_couplers_2in": conduit_couplers_2in,
            "conduit_couplers_4in": conduit_couplers_4in,
            "total_sp1": total_sp1,
            "total_sp2": total_sp2,
            "total_sp3_excluding_f1": total_sp3_excluding_f1,
            "fiber_12": fiber_12,
            "fiber_24": fiber_24,
            "fiber_48": fiber_48,
            "fiber_96": fiber_96,
            "fiber_144": fiber_144,
            "fiber_288": fiber_288,
            "fiber_432": fiber_432,
            "total_heatshrink": total_heatshrink,
            "fp_count": fp_count,
            "sv_count": sv_count,
            "mv_count": mv_count,
            "lv_count": lv_count,
            "xl_count": xl_count,
            "xsv_count": xsv_count,
            "nid_count": nid_count,
            "axl_count": axl_count,
            "coyote_count": coyote_count,
            "x17_count": x17_count,
            "x22_count": x22_count,
            "x28_count": x28_count,
            "x19_count": x19_count,
            "runt_count": runt_count,
            "total_closure_count": total_closure_count,
            "hanger_bracket": hanger_bracket,
            "offset_bracket": offset_bracket,
            "drop_count": drop_count,
            "total_hhp_mdu": total_hhp_mdu,
            "total_dnb_addresses": total_dnb_addresses,
            "slackloop_12_ug": slackloop_12_ug,
            "slackloop_24_ug": slackloop_24_ug,
            "slackloop_48_ug": slackloop_48_ug,
            "slackloop_96_ug": slackloop_96_ug,
            "slackloop_144_ug": slackloop_144_ug,
            "slackloop_288_ug": slackloop_288_ug,
            "pfd_1": pfd_1,
            "total_strand_ftg": total_strand_ftg,
            "est_total_miles": est_total_miles,
            "ae_bom_miles": ae_bom_miles,
            "ug_bom_miles": ug_bom_miles,
            "percent_ae":  percent_ae,
            "percent_ug": percent_ug,
            "lash_closure_count": lash_closure_count,
            "total_f1": total_f1,
            "total_f2": total_f2,
            "total_f1_miles": total_f1_miles,
            "total_f1_ae_miles": total_f1_ae_miles,
            "total_f1_ug_miles": total_f1_ug_miles,
            "total_f2_miles": total_f2_miles,
            "total_f2_ae_miles": total_f2_ae_miles,
            "total_f2_ug_miles": total_f2_ug_miles,
            "total_f2_ug": total_f2_ug,
            "total_f2_ae": total_f2_ae,
            "total_ae_ftg": total_ae_ftg,
            "total_ug_ftg": total_ug_ftg,
            "total_f1_ae": total_f1_ae,
            "total_f1_ug": total_f1_ug,
            "passive_144": passive_144,
            "passive_288": passive_288,
            "passive_432": passive_432,
            "passive_576": passive_576,
            "ug_closure_count": ug_closure_count,
            "snowshoes": snowshoes,
            "pfa_2": pfa_2,
            "total_risers": total_risers,
            "cab_id": cab_id,
            "serv_area": serv_area,
            "city_code": city_code,
            "const_ven": const_ven,
            "total_strand_ftg_reareasment_y": total_strand_ftg_reareasment_y,
            "total_fiber_footage_ug_linear": total_fiber_footage_ug_linear,
            "total_fiber_footage_ae_linear": total_fiber_footage_ae_linear,
            "total_cabinets": total_cabinets,
            "count_over_600ft": count_over_600ft,
            "average_calcfootage": average_calcfootage,
            "total_pole_count": total_pole_count,
            "mr_filtered_pole_count": mr_filtered_pole_count,
            "total_linear_footage": total_linear_footage,
            "active_cabinet_count": active_cabinet_count,
            "grounded_poles": grounded_poles,
            "tree_trimming": tree_trimming,
            "total_ug1ft_reareasment_Y": total_ug1ft_reareasment_Y,
            "e_epmrt_1": e_epmrt_1,
            "total_hhp": total_hhp,
            "down_count": down_count,
            "dirt_count": dirt_count,
            "rock_count": rock_count,
            "total_anchors": total_anchors,
            "uguard_adapter": uguard_adapter,
            "lashing_wire": lashing_wire,
            "special_crossing": special_crossing
        }

    # Exporting to Excel
    run_export = arcpy.GetParameterAsText(1)
    construction_vendor_rate = arcpy.GetParameterAsText(2)  
    design_vendor_rate = arcpy.GetParameterAsText(3)  

    if run_export == "Yes":
        from datetime import datetime
        if not construction_vendor_rate or not design_vendor_rate:
            arcpy.AddWarning("⚠️ Export selected, but Vendors were not provided. Skipping Excel export.\n")
        else:
            # Fallback name
            default_filename = "Exported_BOM.xlsx"
            if cab_id:
                timestamp = datetime.now().strftime("%m-%d-%Y_%H%M%S")
                default_filename = f"BOM_{cab_id}_{timestamp}.xlsx"

            output_path = arcpy.GetParameterAsText(4)

            if not output_path:
                # Build default path in OneDrive
                one_drive_docs = get_one_drive_documents()
                output_path = os.path.join(one_drive_docs, default_filename)
                arcpy.AddMessage(f"No output path specified. Using default: {output_path}")

            # Ensure it ends with .xlsx
            if not output_path.lower().endswith(".xlsx"):
                output_path += ".xlsx"

            # Locate the Excel template inside the "data" folder
            script_dir = os.path.dirname(os.path.abspath(__file__))
            template_path = os.path.join(script_dir, "TEST_BOM_Template.xlsx")

            if not os.path.exists(template_path):
                raise FileNotFoundError(f"Excel template not found: {template_path}")

            arcpy.AddMessage("► Calling export_to_excel function now...")

            # call the primary function for the BOM
            try:
                script_dir = os.path.dirname(os.path.abspath(__file__))
                template_path = os.path.join(script_dir, "TEST_BOM_Template.xlsx")

                if not os.path.exists(template_path):
                    raise FileNotFoundError

            except Exception as e:
                arcpy.AddError(f"❌ Could not find Excel template alongside the script: {e}")
                raise FileNotFoundError("Excel template not found in the expected script folder.")
            export_to_excel(template_path, output_path, values_dict, construction_vendor_rate, design_vendor_rate)

            arcpy.SetParameter(4, output_path)  # ← Make sure param 4 is the output Excel file in your toolbox!

    else:
        output_path = None
        construction_vendor_rate = None
        design_vendor_rate = None
